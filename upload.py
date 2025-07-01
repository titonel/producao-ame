import pandas as pd
from sqlalchemy import create_engine
import openpyxl

# Caminho do arquivo
caminho_arquivo = 'Relatorio_Produção_Executante_01_07_2025_10-29-56.xlsx'

# Lê o conteúdo do Excel a partir da planilha ativa (assume primeira planilha)
wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
ws = wb.active

# Extrai tipo de consulta e data da produção
tipo_consulta = ws['A3'].value
data_producao = ws['F3'].value  # ex: "junho de 2025"
mes_producao, ano_producao = map(str.strip, data_producao.split('de'))

# Lê os dados a partir da linha 7
df = pd.read_excel(caminho_arquivo, skiprows=6)
# Remove colunas de "E" em diante (mantém apenas as 4 primeiras colunas)
df = df.iloc[:, :4]
# Renomeia colunas esperadas (confirme se os nomes estão corretos após leitura)
df.columns = ['Especialidade', 'Oferta', 'Agendados', 'Realizados']

# Remove linhas onde 'Oferta' está vazia ou contém "Total"
df = df[df['Oferta'].notna()]
df = df[df['Oferta'].astype(str).str.lower() != 'total']

# Adiciona colunas auxiliares
df['Tipo_Consulta'] = tipo_consulta
df['Mes_Producao'] = mes_producao
df['Ano_Producao'] = ano_producao

# Criação da conexão SQLite (pode ser adaptada para PostgreSQL, MySQL, etc.)
engine = create_engine('sqlite:///producao.db')

# Grava a tabela no banco
df.to_sql('producao', con=engine, if_exists='replace', index=False)

print("Importação concluída com sucesso.")
