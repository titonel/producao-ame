import pandas as pd
import openpyxl
from io import BytesIO
import os
import streamlit as st # Importado para usar st.warning, st.error, st.success

# --- Funções Auxiliares para Normalização de Especialidades ---
def normalizar_especialidade(nome):
    """Normaliza nomes de especialidades para agrupamento."""
    nome = str(nome).upper().strip()
    if nome.startswith("CIRURGIA PLÁSTICA"):
        return "Cirurgia Plástica"
    elif nome.startswith("CIRURGIA GERAL"):
        return "Cirurgia Geral"    
    elif nome.startswith("CIRURGIA VASCULAR"):
        return "Cirurgia Vascular"    
    elif nome.startswith("CIRURGIA PEDIÁTRICA"):
        return "Cirurgia Pediátrica"    
    elif nome.startswith("OFTALMOLOGIA"):
        return "Oftalmologia"
    elif nome.startswith("DERMATOLOGIA"):
        return "Dermatologia"
    elif nome.startswith("ANESTESIOLOGIA"):
        return "Anestesiologia"
    elif nome.startswith("CARDIOLOGIA"):
        return "Cardiologia"
    elif nome.startswith("COLOPROCTOLOGIA"):
        return "Coloproctologia"
    elif nome.startswith("GASTROCLÍNICA") or nome.startswith("GASTROENTEROLOGIA"):
        return "Gastroenterologia"
    elif nome.startswith("MASTOLOGIA"):
        return "Mastologia"
    elif nome.startswith("ORTOPEDIA"):
        return "Ortopedia"
    elif nome.startswith("OTORRINOLARINGOLOGIA"):
        return "Otorrinolaringologia"    
    elif nome.startswith("UROLOGIA"):
        return "Urologia"
    elif nome.startswith("ENDOCRINOLOGIA"):
        return "Endocrinologia"
    elif nome.startswith("NEUROLOGIA PEDIÁTRICA"):
        return "Neurologia Pediátrica"
    elif nome.startswith("NEUROLOGIA"):
        return "Neurologia Adulto"
    elif nome.startswith("PNEUMOLOGIA PEDIÁTRICA"):
        return "Pneumologia Pediátrica"
    elif nome.startswith("PNEUMOLOGIA"):
        return "Pneumologia"    
    elif nome.startswith("NEFROLOGIA"):
        return "Nefrologia"
    # Adicione outras regras conforme necessário ou retorne o próprio nome se não houver correspondência
    return nome

def process_siresp_upload(uploaded_file_producao, engine):
    """
    Processa o arquivo de upload de dados de produção (SIRESP) e salva no banco de dados.
    Suporta arquivos .xlsx, .xls e .csv.
    """
    try:
        file_extension = os.path.splitext(uploaded_file_producao.name)[1].lower()
        
        df = None
        tipo_consulta = "N/A"
        mes_producao = "N/A"
        ano_producao = "N/A"

        if file_extension in [".xlsx", ".xls"]:
            # Tenta extrair metadados para arquivos .xlsx usando openpyxl
            if file_extension == ".xlsx":
                wb = openpyxl.load_workbook(BytesIO(uploaded_file_producao.read()), data_only=True)
                ws = wb.active
                tipo_consulta_cell = ws['A3'].value
                data_producao_cell = ws['F3'].value
                
                if tipo_consulta_cell:
                    tipo_consulta = str(tipo_consulta_cell)
                else:
                    st.warning("Célula A3 (Tipo de Consulta) vazia. Definindo como 'N/A'.")
                    
                if data_producao_cell and "de" in str(data_producao_cell):
                    mes_producao_str, ano_producao_str = map(str.strip, str(data_producao_cell).split('de'))
                    mes_producao = mes_producao_str.capitalize()
                    ano_producao = ano_producao_str
                else:
                    st.warning(f"Formato de data em F3 '{data_producao_cell}' não reconhecido ou vazio. Mês e Ano de Produção serão 'N/A'.")
                    
                uploaded_file_producao.seek(0) # Reseta o ponteiro após a leitura com openpyxl
            else: # Arquivo .xls
                st.info("Para arquivos .xls, a extração automática de 'Tipo de Consulta', 'Mês' e 'Ano' das células A3 e F3 não é suportada diretamente pelo método atual. Eles serão definidos como 'N/A'.")

            # Lê o dataframe para .xlsx e .xls
            df = pd.read_excel(uploaded_file_producao, skiprows=6)
            df = df.iloc[:, :4]
            df.columns = ['Especialidade', 'Oferta', 'Agendados', 'Realizados']

        elif file_extension == ".csv":
            st.info("Para arquivos .csv, a extração automática de 'Tipo de Consulta', 'Mês' e 'Ano' das células A3 e F3 não é aplicável. Eles serão definidos como 'N/A'. Certifique-se de que o CSV contém as colunas 'Especialidade', 'Oferta', 'Agendados' e 'Realizados' no cabeçalho.")
            df = pd.read_csv(uploaded_file_producao)
            # Para CSV, precisamos garantir que as colunas esperadas existam.
            expected_csv_cols = ['Especialidade', 'Oferta', 'Agendados', 'Realizados']
            if not all(col in df.columns for col in expected_csv_cols):
                st.error(f"❌ Erro: O arquivo CSV não contém as colunas esperadas: {', '.join(expected_csv_cols)}. Por favor, verifique o cabeçalho.")
                return # Retorna para parar a execução da função
            df = df[expected_csv_cols].copy() # Seleciona e reordena as colunas
        
        else:
            st.error("❌ Formato de arquivo não suportado. Por favor, faça o upload de um arquivo .xlsx, .xls ou .csv.")
            return # Retorna para parar a execução da função

        # Processamento comum para todos os tipos de arquivo
        if df is not None:
            # Remove linhas inválidas (ex: somas 'Total')
            df = df[df['Oferta'].notna()]
            df = df[df['Oferta'].astype(str).str.lower() != 'total']

            # Adiciona colunas auxiliares com metadados extraídos/padrão
            df['Tipo_Consulta'] = tipo_consulta
            df['Mes_Producao'] = mes_producao
            df['Ano_Producao'] = ano_producao

            # Salva no banco de dados
            df.to_sql('producao', con=engine, if_exists='append', index=False)

            st.success("✅ Dados de produção inseridos com sucesso!")
            st.subheader("📄 Visualização dos Dados de Produção Inseridos")
            st.dataframe(df)

    except Exception as e:
        st.error(f"❌ Erro ao processar o arquivo de produção: {e}")
        st.exception(e) # Exibe o traceback completo para depuração

def process_contratos_upload(uploaded_file_contratos, engine):
    """
    Processa o arquivo de upload de dados de custos médicos (contratos) e salva no banco de dados.
    """
    try:
        df_contratos = pd.read_excel(uploaded_file_contratos)

        # Renomeia a primeira coluna se for 'Área' para 'Especialidade'
        if df_contratos.columns[0] == 'Área':
            df_contratos.rename(columns={'Área': 'Especialidade'}, inplace=True)
            st.info("A coluna 'Área' foi automaticamente renomeada para 'Especialidade'.")
        
        required_columns = [
            'Especialidade', 'Serviço', 'Centro de Custo', 'Nome do Centro de Custo',
            'Valor Unitário', 'Data Contrato', 'Contratado', 'Meta Mensal',
            'Responsável', 'Detalhamento'
        ]

        # 1. Valida nomes das colunas
        if not all(col in df_contratos.columns for col in required_columns):
            missing_cols = [col for col in required_columns if col not in df_contratos.columns]
            st.error(f"❌ Erro: As seguintes colunas obrigatórias não foram encontradas na planilha: {', '.join(missing_cols)}. Certifique-se de que a primeira coluna seja 'Especialidade' ou 'Área'.")
            return # Retorna para parar a execução da função


        df_contratos = df_contratos[required_columns].copy() # Mantém apenas as colunas necessárias e na ordem

        # 2. Validação e Conversão de Tipos
        errors = []

        # 'Centro de Custo': numérico inteiro de 8 dígitos
        df_contratos['Centro de Custo'] = pd.to_numeric(df_contratos['Centro de Custo'], errors='coerce')
        invalid_cc = df_contratos['Centro de Custo'].isna() | (df_contratos['Centro de Custo'] < 10000000) | (df_contratos['Centro de Custo'] > 99999999) | (df_contratos['Centro de Custo'] % 1 != 0)
        if invalid_cc.any():
            errors.append("Centro de Custo deve ser um número inteiro de 8 dígitos. Verifique as linhas com valores inválidos.")
            df_contratos.loc[invalid_cc, 'Centro de Custo'] = None # Marca como inválido

        # 'Valor Unitário': numérico com 2 casas decimais (float)
        df_contratos['Valor Unitário'] = pd.to_numeric(df_contratos['Valor Unitário'], errors='coerce')
        if df_contratos['Valor Unitário'].isna().any():
            errors.append("Valor Unitário deve ser um número. Verifique as linhas com valores inválidos.")
        
        # 'Data Contrato': formato dd/mm/aaaa
        df_contratos['Data Contrato'] = pd.to_datetime(df_contratos['Data Contrato'], format='%d/%m/%Y', errors='coerce')
        if df_contratos['Data Contrato'].isna().any():
            errors.append("Data Contrato deve estar no formato DD/MM/AAAA. Verifique as linhas com valores inválidos.")
        
        # Outros campos como texto
        for col in ['Especialidade', 'Serviço', 'Nome do Centro de Custo', 'Contratado', 'Meta Mensal', 'Responsável', 'Detalhamento']:
            df_contratos[col] = df_contratos[col].astype(str).replace('nan', '', regex=False).str.strip()


        if errors:
            st.error("❌ Foram encontrados erros de validação na planilha:")
            for err in errors:
                st.write(f"- {err}")
            st.write("Por favor, corrija a planilha e tente novamente.")
            st.dataframe(df_contratos.head()) # Mostra as primeiras linhas para depuração
        else:
            # Tenta criar a tabela se não existir
            from sqlalchemy import text # Importa text aqui para evitar circular import se usado apenas nesta função
            with engine.connect() as connection:
                connection.execute(text("""
                    CREATE TABLE IF NOT EXISTS contratos (
                        Especialidade TEXT,
                        Servico TEXT,
                        "Centro de Custo" INTEGER,
                        "Nome do Centro de Custo" TEXT,
                        "Valor Unitario" REAL,
                        "Data Contrato" DATE,
                        Contratado TEXT,
                        "Meta Mensal" TEXT,
                        Responsavel TEXT,
                        Detalhamento TEXT
                    )
                """))
                connection.commit()

            # Salva no banco de dados
            df_contratos.to_sql('contratos', con=engine, if_exists='append', index=False)
            st.success("✅ Dados dos contratos inseridos com sucesso!")
            st.subheader("📄 Visualização dos Dados dos Contratos Inseridos")
            st.dataframe(df_contratos)

    except Exception as e:
        st.error(f"❌ Erro ao processar o arquivo de contratos: {e}")
        st.exception(e) # Exibe o traceback completo para depuração
